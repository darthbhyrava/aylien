# "hotels", "restaurants", "cars", "airlines"
# domain limited to the above
#
# WILL NOT WORK _ BAD REQUEST



import pandas as pd
import openpyxl
import time
from aylienapiclient import textapi

client = textapi.Client("7a228a93", "bcd1bca1d5c975ede7d49d1b9ce611b0")

wb = openpyxl.load_workbook('21Feb_Pos.xlsx')
ws = wb.active
pred_sent_col = ws['L']
man_sent_col = ws['Q']
text_col = ws['U']
rows = len(pred_sent_col)

correct_p = 0.0
all_rows = []
for row in range(1,rows):
    print row
    time.sleep(0.1)
    towrite = {}
    pred_sent = pred_sent_col[row].value
    man_sent = man_sent_col[row].value
    text = text_col[row].value.replace(';','')
    response = client.AspectBasedSentiment({'domain':'politics','text':text})
    print response
    for aspect in response['aspects']:
        print aspects 
#     towrite['Manual Sentiment'] = man_sent
#     towrite['Aylien_Label'] = aylien_sent
#     towrite['Text'] = text.encode('utf-8')
#     if aylien_sent == 'positive':
#         correct_p+=1.0
#     towrite['Pol_Confidence'] = aylien_conf
#     print towrite
#     all_rows.append(towrite)

# percentage = correct_p/float(len(pred_sent_col))*100
# print percentage
# df = pd.DataFrame(all_rows)
# df.to_csv("21Feb_Pos.csv")
